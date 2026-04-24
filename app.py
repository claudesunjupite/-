from flask import Flask, request, send_file, jsonify
import csv, io, os, re
from datetime import datetime
from collections import Counter
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB

THAI_MONTHS = {
    1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน',
    5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฎาคม', 8: 'สิงหาคม',
    9: 'กันยายน', 10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
}


# ── Styles ────────────────────────────────────────────────────────────────────

def _border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

THIN   = _border('thin')
MEDIUM = _border('medium')

FILL_GRAY   = PatternFill('solid', fgColor='D9D9D9')
FILL_LGRAY  = PatternFill('solid', fgColor='F2F2F2')
FILL_RED    = PatternFill('solid', fgColor='FFCCCC')
FILL_YELLOW = PatternFill('solid', fgColor='FFFF99')
NUM_FMT     = '#,##0.00'


# ── CSV parser ────────────────────────────────────────────────────────────────

def parse_csv_bytes(b: bytes) -> list:
    text = b.decode('utf-8-sig')
    rows = list(csv.reader(io.StringIO(text)))
    out = []
    for row in rows[1:]:
        if len(row) < 17:
            continue
        tax_id   = row[5].strip().lstrip("'")
        inv      = row[7].strip().lstrip("'")
        date_str = row[8].strip()
        plate    = row[9].strip()
        location = row[13].strip()
        try:    before_tax = float(row[14].replace(',', ''))
        except: before_tax = 0.0
        try:    tax = float(row[15].replace(',', ''))
        except: tax = 0.0
        try:    total = float(row[16].replace(',', ''))
        except: total = 0.0
        try:    dt = datetime.fromisoformat(date_str)
        except: dt = None
        out.append(dict(tax_id=tax_id, inv=inv, date_str=date_str, dt=dt,
                        plate=plate, location=location,
                        before_tax=before_tax, tax=tax, total=total))
    return out


# ── PDF parser ────────────────────────────────────────────────────────────────

def _be_to_ce_year(y: int) -> int:
    return y - 543 if y > 2400 else y

def _norm_date_str(s: str) -> str:
    """dd/mm/yyyy or dd/mm/yyyy(BE) → YYYY-MM-DD"""
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s.strip())
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{_be_to_ce_year(y):04d}-{mo:02d}-{d:02d}"
    return s

def parse_pdf_bytes(b: bytes):
    """Return (plate_or_None, list[{date_key, date_str, location, before_tax, tax, total}])"""
    plate = None
    items = []
    try:
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''

                # ── find plate number ──────────────────────────────────────
                if plate is None:
                    for line in text.split('\n'):
                        if 'ทะเบียน' in line:
                            m = re.search(r'([ก-ฮ]{1,3}[ก-ฮ]?\s*\d{3,4})', line)
                            if m:
                                plate = m.group(1).strip()
                                break

                # ── extract toll rows from tables ─────────────────────────
                tables = page.extract_tables() or []
                for table in tables:
                    for row in (table or []):
                        if not row:
                            continue
                        cells = [str(c or '').strip() for c in row]
                        row_text = ' '.join(cells)

                        dm = re.search(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})', row_text)
                        if not dm:
                            continue

                        raw_date = dm.group(1)
                        date_key = _norm_date_str(raw_date)

                        # numeric values (last ones in row)
                        nums = []
                        for cell in reversed(cells):
                            clean = cell.replace(',', '').replace(' ', '')
                            try:
                                nums.append(float(clean))
                            except ValueError:
                                pass

                        if not nums:
                            continue

                        total      = nums[0]
                        tax        = nums[1] if len(nums) > 1 else round(total * 7 / 107, 2)
                        before_tax = nums[2] if len(nums) > 2 else round(total - tax, 2)

                        # location = first non-numeric, non-date, non-short cell
                        location = ''
                        for cell in cells:
                            if (cell and len(cell) > 2
                                    and not re.match(r'^[\d,.\s/\-]+$', cell)
                                    and not re.search(r'\d{1,2}[/\-]\d{1,2}[/\-]\d{4}', cell)):
                                location = cell
                                break

                        items.append(dict(date_key=date_key, date_str=raw_date,
                                         location=location,
                                         before_tax=before_tax, tax=tax, total=total))

    except Exception as e:
        print(f'[PDF] error: {e}')

    return plate, items


# ── helpers ───────────────────────────────────────────────────────────────────

def normalize_plate(p: str) -> str:
    return re.sub(r'\s+', '', p or '')

def fmt_date(dt) -> str:
    if dt is None:
        return ''
    return f'{dt.day:02d}/{dt.month:02d}/{dt.year}'

def _csv_date_key(dt) -> str:
    return dt.strftime('%Y-%m-%d') if dt else ''


# ── Excel sheet builder ───────────────────────────────────────────────────────

def build_sheet(ws, plate: str, records: list, pdf_extras: list, tax_id: str):
    records = sorted(records, key=lambda r: r['dt'] or datetime.min)

    def sc(row, col, val=None, bold=False, size=10, color='000000',
           align='center', fill=None, border=THIN, num_fmt=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Kanit', size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fill:   c.fill   = fill
        if border: c.border = border
        if num_fmt: c.number_format = num_fmt
        return c

    # ── Row 1: plate title ────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    sc(1, 1, f'ทะเบียน {plate}', bold=True, size=13, border=None)
    ws.row_dimensions[1].height = 22

    # ── Row 2: date range + tax id ────────────────────────────────────────
    dates = [r['dt'] for r in records if r['dt']]
    date_range = (f'วันที่ {fmt_date(min(dates))} - {fmt_date(max(dates))}   '
                  f'เลขประจำตัวผู้เสียภาษี {tax_id}') if dates else ''
    ws.merge_cells('A2:H2')
    sc(2, 1, date_range, size=10, border=None)
    ws.row_dimensions[2].height = 16

    # ── Row 3: spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Table-1 headers ────────────────────────────────────────────
    H1 = ['ลำดับ', 'ว.ด.ป.', 'ทะเบียนรถ', 'INV.', 'รายการ', 'ก่อนภาษี', 'ภาษี', 'รวมภาษี']
    for col, h in enumerate(H1, 1):
        sc(4, col, h, bold=True, fill=FILL_GRAY, border=MEDIUM)
    ws.row_dimensions[4].height = 18

    # ── Table-1 data rows ─────────────────────────────────────────────────
    r = 5
    t1_data_start = r
    for i, rec in enumerate(records, 1):
        inv    = rec['inv']
        is_eb  = inv.upper().startswith('EB')
        label  = 'ทางด่วนและรถไฟฟ้า' if is_eb else 'การทางพิเศษฯ'
        color  = 'FF0000' if is_eb else '000000'
        fill   = FILL_RED if is_eb else None

        vals   = [i, fmt_date(rec['dt']), rec['plate'], inv, label,
                  rec['before_tax'], rec['tax'], rec['total']]
        aligns = ['center', 'center', 'center', 'center', 'left',
                  'right', 'right', 'right']

        for col, (v, al) in enumerate(zip(vals, aligns), 1):
            sc(r, col, v, color=color, align=al, fill=fill, border=THIN,
               num_fmt=NUM_FMT if col >= 6 else None)
        ws.row_dimensions[r].height = 15
        r += 1

    # ── Table-1 total row ─────────────────────────────────────────────────
    t1_total = r
    t1_end   = r - 1
    totals1  = ['', '', '', '', 'รวมทั้งหมด',
                f'=SUM(F{t1_data_start}:F{t1_end})',
                f'=SUM(G{t1_data_start}:G{t1_end})',
                f'=SUM(H{t1_data_start}:H{t1_end})']
    for col, v in enumerate(totals1, 1):
        sc(t1_total, col, v, bold=True, fill=FILL_LGRAY, border=THIN,
           align='right' if col >= 5 else 'center',
           num_fmt=NUM_FMT if col >= 6 else None)
    ws.row_dimensions[t1_total].height = 16
    r += 1

    # ── Table-2 (PDF extras) ──────────────────────────────────────────────
    t2_total = None
    if pdf_extras:
        r += 3  # 3 blank rows

        H2 = ['ลำดับ', 'ว.ด.ป.', 'ทะเบียนรถ', 'รายการ',
               'ก่อนภาษี', 'ภาษี', 'รวมภาษี']
        for col, h in enumerate(H2, 1):
            sc(r, col, h, bold=True, fill=FILL_GRAY, border=MEDIUM)
        ws.row_dimensions[r].height = 18
        r += 1

        t2_data_start = r
        for i, item in enumerate(pdf_extras, 1):
            vals2   = [i, item['date_str'], plate, item['location'],
                       item['before_tax'], item['tax'], item['total']]
            aligns2 = ['center', 'center', 'center', 'left',
                       'right', 'right', 'right']
            for col, (v, al) in enumerate(zip(vals2, aligns2), 1):
                sc(r, col, v, align=al, border=THIN,
                   num_fmt=NUM_FMT if col >= 5 else None)
            ws.row_dimensions[r].height = 15
            r += 1

        t2_total = r
        t2_end   = r - 1
        totals2  = ['', '', '', 'รวมทั้งหมด',
                    f'=SUM(E{t2_data_start}:E{t2_end})',
                    f'=SUM(F{t2_data_start}:F{t2_end})',
                    f'=SUM(G{t2_data_start}:G{t2_end})']
        for col, v in enumerate(totals2, 1):
            sc(t2_total, col, v, bold=True, fill=FILL_LGRAY, border=THIN,
               align='right' if col >= 4 else 'center',
               num_fmt=NUM_FMT if col >= 5 else None)
        ws.row_dimensions[t2_total].height = 16
        r += 1

        # ── Grand total (yellow) ──────────────────────────────────────────
        r += 1  # 1 blank row
        grand = r
        gvals = ['', '', '', 'ยอดรวมทั้งสิ้น',
                 f'=F{t1_total}+E{t2_total}',
                 f'=G{t1_total}+F{t2_total}',
                 f'=H{t1_total}+G{t2_total}']
        for col, v in enumerate(gvals, 1):
            sc(grand, col, v, bold=True, size=11, fill=FILL_YELLOW, border=THIN,
               align='right' if col >= 4 else 'center',
               num_fmt=NUM_FMT if col >= 5 else None)
        ws.row_dimensions[grand].height = 20

    # ── Column widths & freeze ────────────────────────────────────────────
    for col, w in enumerate([8, 13, 14, 24, 22, 14, 11, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A5'


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_file('index.html')


@app.route('/process', methods=['POST'])
def process():
    csv_files = request.files.getlist('csv_files')
    pdf_files = request.files.getlist('pdf_files')

    if not csv_files or all(f.filename == '' for f in csv_files):
        return jsonify({'error': 'กรุณาอัปโหลดไฟล์ CSV อย่างน้อย 1 ไฟล์'}), 400

    # ── Parse CSVs ────────────────────────────────────────────────────────
    plate_data = {}  # norm_plate -> {plate, tax_id, records}
    for f in csv_files:
        if not f.filename:
            continue
        records = parse_csv_bytes(f.read())
        if not records:
            continue
        plate_display = records[0]['plate']
        tax_id        = records[0]['tax_id'] or '0994000165421'
        key           = normalize_plate(plate_display)
        if key not in plate_data:
            plate_data[key] = {'plate': plate_display, 'tax_id': tax_id, 'records': []}
        plate_data[key]['records'].extend(records)

    if not plate_data:
        return jsonify({'error': 'ไม่พบข้อมูลในไฟล์ CSV'}), 400

    # ── Parse PDFs ────────────────────────────────────────────────────────
    pdf_by_plate = {}  # norm_plate -> [items]
    for f in pdf_files:
        if not f.filename:
            continue
        pdf_plate, items = parse_pdf_bytes(f.read())
        if pdf_plate and items:
            key = normalize_plate(pdf_plate)
            pdf_by_plate.setdefault(key, []).extend(items)

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_dates = []
    for pdata in plate_data.values():
        all_dates += [r['dt'] for r in pdata['records'] if r['dt']]

    for key, pdata in plate_data.items():
        records       = pdata['records']
        plate_display = pdata['plate']
        tax_id        = pdata['tax_id']

        # build CSV date-amount counter for matching
        csv_counter = Counter()
        for rec in records:
            if rec['dt']:
                csv_counter[(_csv_date_key(rec['dt']), round(rec['total'], 2))] += 1

        # find PDF items not matched in CSV → Table 2
        pdf_extras = []
        if key in pdf_by_plate:
            ctr = dict(csv_counter)
            for item in pdf_by_plate[key]:
                k = (item['date_key'], round(item['total'], 2))
                if ctr.get(k, 0) > 0:
                    ctr[k] -= 1
                else:
                    pdf_extras.append(item)

        sheet_name = re.sub(r'[\\/*?:\[\]]', '', plate_display)[:31]
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, plate_display, records, pdf_extras, tax_id)

    # ── Filename ──────────────────────────────────────────────────────────
    thai_month = THAI_MONTHS[max(all_dates).month] if all_dates else 'รวม'
    filename   = f'รวม{thai_month}.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''{quote(filename)}"
    )
    return response


if __name__ == '__main__':
    app.run(debug=True, port=5000)
